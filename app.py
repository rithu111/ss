from flask import Flask, render_template, request, redirect, url_for, flash, session
import secrets
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key =secrets.token_hex(16)
# Load the existing Excel file
excel_file = 'stud.xlsx'
wb = load_workbook(excel_file)
ws = wb.active
@app.route('/')

def index():
    return render_template('index.html')
@app.route('/add', methods=['GET', 'POST'])
def add_student():
    if request.method == 'POST':
      
      
       name = request.form.get('name')
       age = request.form.get('age')
       email = request.form.get('email')

        # Find the next available row and append data
       next_row = ws.max_row + 1
       ws.cell(row=next_row, column=1, value=name)
       ws.cell(row=next_row, column=2, value=age)
       ws.cell(row=next_row, column=3, value=email)
        
        # Save the workbook
       wb.save(excel_file)

        
    return render_template('add.html')
@app.route('/add2')
def add2():
    return render_template('add2.html')

@app.route('/search', methods=['GET','POST'])
def search():
    if request.method == 'POST':
        search_name = request.form['search_name']
        found = False

        for row in ws.iter_rows(values_only=True):
            if row[0] == search_name:
                found = True
                return f"Name: {row[0]}\nAge: {row[1]}\nEmail: {row[2]}"
        if not found:
            return "Information not found."
             
    return render_template('search.html')
@app.route('/edit', methods=['GET', 'POST'])
def edit_student():
    if request.method == 'POST':
        search_name = request.form['search_name']
        found = False

        for row in ws.iter_rows(values_only=True):
            if row[0] == search_name:
                found = True
                original_name = row[0]
                return render_template('edit2.html', search_name=search_name, found=found, original_name=original_name)

        if not found:
            return "Information not found."

    return render_template('edit.html')

@app.route('/edit2', methods=['POST'])
def perform_edit2():
    found = request.form.get('found')
    original_name = request.form.get('original_name')
    
    if found and found == 'True':
        new_age = request.form['new_age']
        new_email = request.form['new_email']

        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == original_name:
                ws.cell(row=row_index, column=2, value=new_age)
                ws.cell(row=row_index, column=3, value=new_email)
                wb.save(excel_file)
                return "Student details edited successfully.<a href='/'>Go to Home</a>"

    return "Information not found."

    return render_template('edit2.html')

@app.route('/delete', methods=['GET', 'POST'])
def delete():
    return render_template('delete.html')
@app.route('/delete2', methods=['GET','POST'])
def delete2():
    confirm_delete = request.form.get('confirm_delete')
    found = False  # Initialize found with False
    
    if confirm_delete:
        search_name = request.form.get('search_name')

        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if str(row[0]) == search_name:
                found = True
                ws.delete_rows(row_index)
                wb.save(excel_file)
                return f"Student '{search_name}' deleted successfully."

        if not found:
            return f"Student '{search_name}' not found."

    
    return render_template('delete2.html')


if __name__ == '__main__':
    app.run(debug=True)
